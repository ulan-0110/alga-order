import streamlit as st
import pandas as pd
import openpyxl
import json
import os
import io
import traceback
import re
from datetime import datetime

st.set_page_config(page_title="SmartOrder Web — АлгаДистрибьюшн", page_icon="📊", layout="wide")

# ==========================================
# СИСТЕМА ЛОГИРОВАНИЯ И ОЧИСТКИ ДАННЫХ
# ==========================================
def log_error(context, exception):
    log_filename = "error_log.txt"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    error_msg = f"[{timestamp}] КРИТИЧЕСКАЯ ОШИБКА в блоке: {context}\nОписание: {str(exception)}\nТрассировка кода:\n{traceback.format_exc()}\n{'='*80}\n\n"
    with open(log_filename, "a", encoding="utf-8") as f:
        f.write(error_msg)

def parse_number(val, num_type=float):
    """Вытаскивает чистые цифры из любой ячейки (спасает от пробелов, долларов и запятых)"""
    if val is None or pd.isna(val): return num_type(0)
    s = str(val).replace(",", ".").replace(" ", "").strip()
    s = re.sub(r'[^\d\.-]', '', s)
    try: return num_type(s) if s else num_type(0)
    except: return num_type(0)

def normalize_strict(name):
    """ЖЕЛЕЗОБЕТОННАЯ нормализация. Оставляет только буквы и цифры.
    Убирает кавычки, пробелы и елочки для точного сравнения баз."""
    if not isinstance(name, str) or pd.isna(name): return ""
    return re.sub(r'[^\w\d]', '', str(name).lower())

# ==========================================
# БЛОК БЕЗОПАСНОСТИ (3 РАЗДЕЛЬНЫХ ДОСТУПА)
# ==========================================
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False
        st.session_state["username"] = "" # Храним имя текущего пользователя
        
    if st.session_state["authenticated"]: return True

    st.title("🔒 Вход в систему Alga Distribution")
    login_input = st.text_input("Логин", key="login_field").strip().lower()
    password_input = st.text_input("Пароль", type="password", key="password_field")
    
    if st.button("Войти"):
        # Карта пользователей: Логин -> Пароль
        # Сюда ты можешь вписать любые свои 3 логина и пароля
        users_credentials = {
            "ulan": "alga2026",
            "osh": "alga_osh",
            "bishkek": "alga_bishkek"
        }
        
        # Если в облаке в Secrets прописаны кастомные доступы, они заменят дефолтные
        if hasattr(st, "secrets"):
            for k, v in st.secrets.items():
                if k.startswith("USER_"):
                    # Формат в Secrets: USER_ulan = "пароль"
                    u_name = k.replace("USER_", "").lower()
                    users_credentials[u_name] = str(v)

        if login_input in users_credentials and password_input == users_credentials[login_input]:
            st.session_state["authenticated"] = True
            st.session_state["username"] = login_input # Запоминаем, кто именно вошел!
            st.rerun()
        else:
            st.error("❌ Неверный логин или пароль!")
    return False

if not check_password(): st.stop()

# Берем имя пользователя для названия индивидуального файла бэкапа
current_user = st.session_state["username"]
backup_filename = f"{current_user}_order_state.json"

# ==========================================
# ОБЛАЧНЫЕ ФУНКЦИИ БАЗЫ ЗНАНИЙ
# ==========================================
@st.cache_data
def load_mapping_cloud():
    if os.path.exists("mapping.csv"):
        try: return pd.read_csv("mapping.csv", sep=";")
        except: pass
    return pd.DataFrame(columns=["Номенклатура АлгаДистрибьюшн факт", "Наименование от производителя"])

def save_new_pair_cloud(factory_name, our_name):
    try:
        mapping_df = load_mapping_cloud()
        new_row = pd.DataFrame([{"Номенклатура АлгаДистрибьюшн факт": our_name.strip(), "Наименование от производителя": factory_name.strip()}])
        mapping_df = pd.concat([mapping_df, new_row], ignore_index=True)
        mapping_df.to_csv("mapping.csv", sep=";", index=False, encoding="utf-8")
        st.cache_data.clear() 
    except Exception as e: log_error("save_new_pair", e)

# ==========================================
# ОБНОВЛЕННЫЙ ПАРСЕР ЦЕН ШАБЛОНА (БЕЗУСЛОВНЫЙ)
# ==========================================
def check_and_cache_template():
    """Собирает ВСЕ строки, наследует любую найденную цену строго сверху вниз"""
    if not os.path.exists("template.xlsx"): return
    try:
        wb = openpyxl.load_workbook("template.xlsx", data_only=True)
        cache_data = {"products": []}

        if "заказ-order" in wb.sheetnames:
            ws = wb["заказ-order"]
            current_price = 0.0

            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row or len(row) < 7: continue
                name = str(row[1]).strip() if row[1] else ""
                
                if not name: continue

                parsed_price = parse_number(row[6], float)
                if parsed_price > 0:
                    current_price = parsed_price 
                
                item_price = parsed_price if parsed_price > 0 else current_price
                box_size = parse_number(row[2], int)

                cache_data["products"].append({
                    "factory_name": name,
                    "norm_name": normalize_strict(name),
                    "box_size": box_size,
                    "price": item_price
                })

        with open("template_cache.json", "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        check_and_cache_template()

# ==========================================
# ИНТЕРФЕЙС
# ==========================================
col_t1, col_t2 = st.columns([9, 1])
with col_t1: st.title(f"📊 SmartOrder Web — Панель [{current_user.upper()}]")
with col_t2:
    if st.button("🚪 Выйти"):
        st.session_state["authenticated"] = False
        st.session_state["username"] = ""
        st.rerun()

tab1, tab2 = st.tabs(["📥 Расшифровка входящего заказа", "📤 Набрать новый заказ для завода"])

# ------------------------------------------
# ВКЛАДКА 1: ОБРАБОТКА ПОСТУПИВШЕГО ЗАКАЗА
# ------------------------------------------
with tab1:
    st.header("Загрузка и очистка файла от завода")
    uploaded_file = st.file_uploader("Перетащите сюда Excel или CSV файл заказа", type=["xlsx", "xls", "csv"])
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'): df_raw = pd.read_csv(uploaded_file, header=None, sep=";")
            else: df_raw = pd.read_excel(uploaded_file, header=None)
                
            mapping_df = load_mapping_cloud()
            existing_1c_items = sorted(mapping_df["Номенклатура АлгаДистрибьюшн факт"].dropna().unique().tolist())
            
            if "[ИГНОР_КАТЕГОРИЯ]" in existing_1c_items:
                existing_1c_items.remove("[ИГНОР_КАТЕГОРИЯ]")
                
            mapping_dict = {normalize_strict(row["Наименование от производителя"]): str(row["Номенклатура АлгаДистрибьюшн факт"]).strip() for _, row in mapping_df.iterrows()}
            
            current_price = 0.0
            processed_rows = []
            unknown_products = []
            
            for idx, row in df_raw.iloc[4:].iterrows():
                if len(row) < 7: continue
                name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                
                name_lower = name.lower()
                if not name or "всего" in name_lower or "итого" in name_lower or "общая стоимость" in name_lower: 
                    continue

                box_size = parse_number(row.iloc[2] if len(row) > 2 else 0, int)
                boxes = parse_number(row.iloc[4] if len(row) > 4 else 0, int)
                parsed_price = parse_number(row.iloc[6] if len(row) > 6 else None, float)

                if parsed_price > 0:
                    current_price = parsed_price 
                
                factory_norm_name = normalize_strict(name)
                our_name = mapping_dict.get(factory_norm_name, "")
                
                if our_name == "[ИГНОР_КАТЕГОРИЯ]":
                    continue
                
                if box_size == 0 or boxes == 0:
                    continue 

                pcs = parse_number(row.iloc[3] if len(row) > 3 else 0, int)
                item_price = parsed_price if parsed_price > 0 else current_price
                    
                if not our_name:
                    if not any(p['norm_name'] == factory_norm_name for p in unknown_products):
                        unknown_products.append({"name": name, "norm_name": factory_norm_name, "row_num": idx + 1})
                        
                display_1c_name = our_name if our_name else f"[НЕОПРЕДЕЛЕН] {name}"
                processed_rows.append({
                    "Ваша Номенклатура (1С)": display_1c_name, 
                    "Наименование Завода": name, 
                    "Ящиков": boxes, 
                    "Штук в ящ": box_size, 
                    "Всего штук": pcs, 
                    "Цена ($)": item_price, 
                    "Сумма ($)": pcs * item_price
                })
            
            if unknown_products:
                st.error(f"⚠️ Найдено {len(unknown_products)} позиций в заказе, которых нет в базе 1С.")
                for i, prod_info in enumerate(unknown_products[:3]):
                    st.markdown("---")
                    st.info(f"🏭 В файле заказа строка {prod_info['row_num']}: `{prod_info['name']}`")
                    
                    category_option = "-- 🛑 ЭТО КАТЕГОРИЯ (Пропустить и запомнить как заголовок) --"
                    options_list = [category_option, "-- Выбрать из существующих в 1С --"] + existing_1c_items
                    
                    selected_1c_name = st.selectbox("Что это за позиция?:", options=options_list, key=f"sel_{i}")
                    manual_1c_name = st.text_input("Или ввести новое имя 1С вручную (если это реальный товар):", key=f"man_{i}")
                    
                    if st.button("Зафиксировать выбор", key=f"btn_{i}"):
                        if selected_1c_name == category_option:
                            final_1c_name = "[ИГНОР_КАТЕГОРИЯ]"
                        else:
                            final_1c_name = manual_1c_name.strip() if manual_1c_name.strip() else (selected_1c_name if selected_1c_name != "-- Выбрать из существующих в 1С --" else "")
                        
                        if final_1c_name:
                            save_new_pair_cloud(prod_info['name'], final_1c_name)
                            st.success(f"✅ Выбор зафиксирован!")
                            st.rerun()
                st.stop()
            else:
                if processed_rows:
                    df_res = pd.DataFrame(processed_rows)
                    st.markdown("### 📋 Итоги распознавания заказа:")
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Распознано SKU", len(df_res))
                    c2.metric("Всего коробок", int(df_res["Ящиков"].sum()))
                    c3.metric("Общая сумма", f'{df_res["Сумма ($)"].sum():,.2f} $')
                    
                    st.dataframe(df_res, use_container_width=True)
                    
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer: 
                        df_res.to_excel(writer, index=False)
                    st.download_button(label="💾 Скачать очищенный Excel для 1С", data=buffer.getvalue(), file_name="Обработанный_Заказ_АлгаWeb.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            log_error("Вкладка1_Парсинг", e)
            st.error("Ошибка обработки файла.")

# ------------------------------------------
# ВКЛАДКА 2: НАБОР НОВОГО ЗАКАЗА ДЛЯ ЗАВОДА
# ------------------------------------------
with tab2:
    st.header("Единый бланк заказа")
    st.write("Скролл зафиксирован. Заполняйте ящики в таблице — автосохранение работает раздельно для каждого менеджера.")
    
    if st.button("🔄 Сбросить и обновить справочник цен из template.xlsx"):
        with st.spinner("Перечитываем и пересчитываем файл шаблона..."):
            check_and_cache_template()
            st.cache_data.clear() 
            if os.path.exists(backup_filename):
                os.remove(backup_filename) # Чистим сейв ТЕКУЩЕГО пользователя при сбросе
            if "last_processed_file_key" in st.session_state:
                del st.session_state["last_processed_file_key"]
            st.success("База цен и связок успешно обновлена!")
            st.rerun()

    if not os.path.exists("template_cache.json"):
        with st.spinner("Собираем базу данных..."):
            check_and_cache_template()
            
    if os.path.exists("template_cache.json"):
        try:
            with open("template_cache.json", "r", encoding="utf-8") as f:
                cache = json.load(f)
            
            factory_cache_dict = {p["norm_name"]: {"box_size": p["box_size"], "price": p["price"]} for p in cache["products"]}
                
            mapping_df = load_mapping_cloud()
            mapping_clean = mapping_df.dropna(subset=["Номенклатура АлгаДистрибьюшн факт", "Наименование от производителя"])
            mapping_clean = mapping_clean[mapping_clean["Номенклатура АлгаДистрибьюшн факт"] != "[ИГНОР_КАТЕГОРИЯ]"]
            mapping_clean = mapping_clean.drop_duplicates(subset=["Номенклатура АлгаДистрибьюшн факт"])
            mapping_clean = mapping_clean.sort_values(by="Номенклатура АлгаДистрибьюшн факт")
            
            # ==========================================
            # АВТОМАТИЧЕСКАЯ ЗАГРУЗКА ЗАКАЗА ИЗ 1С
            # ==========================================
            st.markdown("### 📥 Автоматическое заполнение из файла 1С")
            tab2_uploaded_file = st.file_uploader("Перетащите сюда Excel или CSV файл с заказом 1С", type=["xlsx", "xls", "csv"], key="tab2_file_uploader")
            
            # ЗАГРУЗКА ИНДИВИДУАЛЬНОГО АВТОСОХРАНЕНИЯ МЕНЕДЖЕРА
            saved_boxes_dict = {}
            if os.path.exists(backup_filename):
                try:
                    with open(backup_filename, "r", encoding="utf-8") as sf:
                        saved_boxes_dict = json.load(sf)
                except:
                    pass

            if tab2_uploaded_file is not None:
                file_key = f"processed_{tab2_uploaded_file.name}_{tab2_uploaded_file.size}"
                
                # Если файл еще не применен, показываем кнопку
                if st.session_state.get("last_processed_file_key") != file_key:
                    st.info("📄 Файл прочитан. Нажмите кнопку ниже, чтобы применить данные к корзине (старая корзина будет очищена).")
                    
                    if st.button("📥 Сформировать корзину из файла", type="primary"):
                        try:
                            if tab2_uploaded_file.name.endswith('.csv'):
                                df_up = pd.read_csv(tab2_uploaded_file, header=None, sep=";")
                            else:
                                df_up = pd.read_excel(tab2_uploaded_file, header=None)
                            
                            our_1c_norm_dict = {normalize_strict(row["Номенклатура АлгаДистрибьюшн факт"]): str(row["Номенклатура АлгаДистрибьюшн факт"]).strip() for _, row in mapping_clean.iterrows()}
                            
                            # ЖЕСТКАЯ ОЧИСТКА КОРЗИНЫ ПЕРЕД ИМПОРТОМ НОВОГО ФАЙЛА
                            saved_boxes_dict.clear()
                            matched_count = 0
                            
                            for _, r in df_up.iterrows():
                                if len(r) < 2: continue
                                u_name = str(r.iloc[0]).strip()
                                if not u_name or u_name.lower() in ["наименование", "товар", "итого", "всего", "номенклатура", "nan"]: 
                                    continue
                                
                                u_norm = normalize_strict(u_name)
                                if u_norm in our_1c_norm_dict:
                                    exact_1c_name = our_1c_norm_dict[u_norm]
                                    u_boxes = parse_number(r.iloc[1] if len(r) > 1 else 0, int)
                                    u_pcs = parse_number(r.iloc[2] if len(r) > 2 else 0, int)
                                    
                                    # Если в файле 1С будут дубликаты строк, мы их плюсуем
                                    if u_boxes > 0:
                                        saved_boxes_dict[exact_1c_name] = saved_boxes_dict.get(exact_1c_name, 0) + u_boxes
                                        matched_count += 1
                                    elif u_pcs > 0:
                                        f_name = str(mapping_clean[mapping_clean["Номенклатура АлгаДистрибьюшн факт"] == exact_1c_name]["Наименование от производителя"].values[0]).strip()
                                        f_name_key = normalize_strict(f_name)
                                        factory_info = factory_cache_dict.get(f_name_key, {"box_size": 0, "price": 0.0})
                                        b_size = int(factory_info["box_size"])
                                        if b_size > 0:
                                            saved_boxes_dict[exact_1c_name] = saved_boxes_dict.get(exact_1c_name, 0) + int(u_pcs // b_size)
                                            matched_count += 1
                            
                            # 1. Сохраняем результат в файл бэкапа
                            with open(backup_filename, "w", encoding="utf-8") as sf:
                                json.dump(saved_boxes_dict, sf, ensure_ascii=False)
                            
                            # 2. Сбрасываем кэш таблицы, чтобы виджет перерисовался с новыми данными
                            if "super_stable_editor" in st.session_state:
                                del st.session_state["super_stable_editor"]
                            
                            st.session_state["last_processed_file_key"] = file_key
                            
                            if matched_count > 0:
                                st.success(f"✅ Файл 1С успешно обработан! Распознано SKU: {matched_count}")
                            else:
                                st.error("⚠️ Внимание: Ни одно наименование из файла не совпало с базой 1С. Проверь названия в загруженном документе.")
                                
                        except Exception as e:
                            log_error("Вкладка2_Автозаполнение_1С", e)
                            st.error("Ошибка при разборе загруженного файла 1С.")

            # ==========================================
            # ОТРИСОВКА ТАБЛИЦЫ БЛАНКА
            # ==========================================
            table_rows = []
            for _, row in mapping_clean.iterrows():
                our_name = str(row["Номенклатура АлгаДистрибьюшн факт"]).strip()
                f_name = str(row["Наименование от производителя"]).strip()
                f_name_key = normalize_strict(f_name)
                
                factory_info = factory_cache_dict.get(f_name_key, {"box_size": 0, "price": 0.0})
                
                # Подтягиваем предзаполненное количество (из файла или старого ручного ввода)
                default_boxes = int(saved_boxes_dict.get(our_name, 0))
                
                table_rows.append({
                    "Номенклатура (1С)": our_name,
                    "Товар завода": f_name,
                    "В коробке (шт)": int(factory_info["box_size"]),
                    "Цена ($)": float(factory_info["price"]),
                    "Заказ (Ящиков)": default_boxes
                })
                
            df_form = pd.DataFrame(table_rows)
            
            if df_form.empty:
                st.warning("⚠️ В файле mapping.csv пока нет связей.")
                st.stop()
            
            # ТАБЛИЦА ВВОДА (Отображается уже заполненной на основе файла)
            edited_df = st.data_editor(
                df_form,
                column_config={
                    "Номенклатура (1С)": st.column_config.TextColumn(disabled=True, width="large"),
                    "Товар завода": st.column_config.TextColumn(disabled=True, width="medium"),
                    "В коробке (шт)": st.column_config.NumberColumn(disabled=True, width="small"),
                    "Цена ($)": st.column_config.NumberColumn(disabled=True, format="$ %.2f", width="small"),
                    "Заказ (Ящиков)": st.column_config.NumberColumn(min_value=0, max_value=5000, step=1, required=True, width="small")
                },
                disabled=["Номенклатура (1С)", "Товар завода", "В коробке (шт)", "Цена ($)"],
                use_container_width=True,
                key="super_stable_editor", 
                hide_index=True
            )

            # ДВИЖОК ИНДИВИДУАЛЬНОГО АВТОСОХРАНЕНИЯ РУЧНЫХ КОРРЕКЦИЙ
            current_boxes_state = {row["Номенклатура (1С)"]: int(row["Заказ (Ящиков)"]) for _, row in edited_df.iterrows()}
            with open(backup_filename, "w", encoding="utf-8") as sf:
                json.dump(current_boxes_state, sf, ensure_ascii=False)

            # МГНОВЕННЫЙ РАСЧЕТ ИТОГОВ КОРЗИНЫ
            edited_df["Заказ (Ящиков)"] = pd.to_numeric(edited_df["Заказ (Ящиков)"]).fillna(0).astype(int)
            edited_df["В коробке (шт)"] = pd.to_numeric(edited_df["В коробке (шт)"]).fillna(0).astype(int)
            edited_df["Цена ($)"] = pd.to_numeric(edited_df["Цена ($)"]).fillna(0.0).astype(float)
            
            edited_df["Итого штук"] = edited_df["Заказ (Ящиков)"] * edited_df["В коробке (шт)"]
            edited_df["Итого сумма ($)"] = edited_df["Итого штук"] * edited_df["Цена ($)"]
            
            active_orders = edited_df[edited_df["Заказ (Ящиков)"] > 0]

            st.markdown("---")
            st.subheader("🛒 Ваша корзина (формируется автоматически):")

            if not active_orders.empty:
                c_m1, c_m2, c_m3 = st.columns(3)
                c_m1.metric("Позиций (SKU)", len(active_orders))
                c_m2.metric("Всего ящиков", int(active_orders["Заказ (Ящиков)"].sum()))
                c_m3.metric("Общая сумма ($)", f'{active_orders["Итого сумма ($)"].sum():,.2f} $')
                
                st.dataframe(
                    active_orders[["Номенклатура (1С)", "Заказ (Ящиков)", "Итого штук", "Цена ($)", "Итого сумма ($)"]], 
                    use_container_width=True, 
                    hide_index=True
                )
                
                if st.button("🚀 Подготовить и скачать файл для завода", type="primary"):
                    if not os.path.exists("template.xlsx"):
                        st.error("Отсутствует оригинальный шаблон template.xlsx")
                    else:
                        with st.spinner("Создаем Excel-файл..."):
                            try:
                                input_data = {normalize_strict(k): v for k, v in zip(active_orders["Товар завода"], active_orders["Заказ (Ящиков)"])}
                                wb = openpyxl.load_workbook("template.xlsx", data_only=False)
                                
                                if "заказ-order" in wb.sheetnames:
                                    ws = wb["заказ-order"]
                                    for row_idx in range(5, ws.max_row + 1):
                                        name_val = ws.cell(row=row_idx, column=2).value
                                        if name_val is None: continue
                                        
                                        f_name_key = normalize_strict(name_val)
                                        if f_name_key in input_data:
                                            boxes_to_order = input_data[f_name_key]
                                            if boxes_to_order > 0:
                                                ws.cell(row=row_idx, column=5).value = boxes_to_order
                                            else:
                                                ws.cell(row=row_idx, column=5).value = None
                                                
                                out_buf = io.BytesIO()
                                wb.save(out_buf)
                                out_buf.seek(0)
                                st.session_state["excel_ready_bytes"] = out_buf.getvalue()
                                
                                # Автоматически стираем личный бэкап и сбрасываем триггер файла, так как заказ успешно собран и скачан!
                                if os.path.exists(backup_filename):
                                    os.remove(backup_filename)
                                if "last_processed_file_key" in st.session_state:
                                    del st.session_state["last_processed_file_key"]
                            except Exception as e:
                                log_error("Генерация_Завода", e)
                                st.error("Ошибка записи данных в Excel.")

                if st.session_state.get("excel_ready_bytes"):
                    st.success("🎉 Файл успешно сгенерирован!")
                    st.download_button(
                        label="📥 СКАЧАТЬ ГОТОВЫЙ ФАЙЛ ДЛЯ ЗАВОДА",
                        data=st.session_state["excel_ready_bytes"],
                        file_name=f"Заказ_Завод_{current_user.upper()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.info("Корзина пуста. Начните вводить количество ящиков или загрузите файл 1С выше.")
                if "excel_ready_bytes" in st.session_state:
                    st.session_state["excel_ready_bytes"] = None
                    
        except Exception as e:
            log_error("Вкладка2_Рендеринг", e)
            st.error("Произошла ошибка при отрисовке бланка заказа.")
